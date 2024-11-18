import numpy as np
import openpyxl

def CarSelect():
    print('****************************************************')
    print('             CAR PERFORMANCE ANALYSIS               ')
    print('****************************************************')
    print('Choose the vehicle type: ')
    print('1.  2011 Mazda RX-8                  Rotary')
    print('2.  2011 Ford F-250                  Diesel')
    print('3.  2004 Chevrolet Corvette          NA V8')
    print('4.  2015 GT-R                        Turbo V6')
    print('5.  2015 Audi S4                     Supercharged V6')
    print('****************************************************')
    print('6. EXIT')
    print('****************************************************')
    carSelections = []
    while(6 not in carSelections):
            carSelection = float(input('Select a car type (1-5) or EXIT (6): '))
            match carSelection:
                case 1:
                    carSelections.append(carSelection)
                case 2:
                    carSelections.append(carSelection)
                case 3:
                    carSelections.append(carSelection)
                case 4:
                    carSelections.append(carSelection)
                case 5:
                    carSelections.append(carSelection)
                case 6:
                    carSelections.append(carSelection)
                case _:
                    print('ERROR! Please select an appropriate option')
    if carSelections[0] == 6:
        print('****************************************************')
        print('      NO CAR SELECTED. NO ANALYSIS PERFORMED        ')
        print('****************************************************')
    carSelections.pop()
    return carSelections

def ImportData(carNum):
    car = ''
    match carNum:
        case 1:
            car = 'MAZDA'
        case 2:
            car = 'FORD'
        
        case 3:
            car = 'CORVETTE'
            
        case 4:
            car = 'GTR'
            
        case 5:
            car = 'AUDI'

    path = car + '.xlsx'
    workbook = openpyxl.load_workbook(path)
    
    sheet = workbook['DYNO']
    rpm = []
    torque = []
    for cell in sheet['A'][1:]:
        rpm.append(cell.value)
    for cell in sheet['B'][1:]:
        torque.append(cell.value)
    
    sheet = workbook['DRIVETRAIN']

    gear = []
    ratio = []

    for cell in sheet['A'][1:]:
        gear.append(cell.value)
    for cell in sheet['B'][1:]:
        ratio.append(cell.value)
    
    i = 2
    for each_cellA, each_cellB in zip(gear,ratio):
        while(each_cellB < 0):
            sheet.cell(i,2,float(input(each_cellA + ' is negative. Please insert it again: ')))            
        i += 1
    workbook.save(car + '.xlsx')

    drivetype = sheet["D2"].value
        




  
    

    return rpm, torque, ratio

def otherInfo():
    workbook = openpyxl.load_workbook("Other_info.xslx")
    sheet = workbook['Sheet1']
    data = sheet["C2:C15"]
    names = sheet["B2:B15"]
    i = 2
    for each_cellA, each_cellB in zip(names,data):
        while(each_cellB < 0):
            sheet.cell(i,2,float(input(each_cellA + ' is negative. Please insert it again: ')))            
        i += 1

    workbook.save('Other_info.xlsx')
    data = sheet["C2:C15"]
    v = sheet["C2"].value
    Gp = sheet["C3"].value
    L = sheet["C4"].value
    r = sheet["C5"].value
    f = sheet["C6"].value
    hAh = sheet["C7"].value
    nd = sheet["C8"].value
    nt = sheet["C9"].value
    W = sheet["C10"].value
    p = sheet["C11"].value 
    id = sheet["C12"].value
    b = sheet["C13"].value
    A = sheet["C14"].value
    cd = sheet["C15"].value
    return v, Gp, L, r, f, hAh, nd, nt, W, p, id, b, A, cd
